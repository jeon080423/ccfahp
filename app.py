import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy import linalg, stats
import io
import warnings
from datetime import datetime

from openpyxl.styles import numbers
from openpyxl.chart import LineChart, Reference

warnings.filterwarnings("ignore")

st.set_page_config(page_title="Fuzzy AHP 분석 시스템", layout="wide", page_icon="📊")

# -----------------------------
# 0. 세션 상태 초기화 (로그인 관련)
# -----------------------------
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "last_login" not in st.session_state:
    st.session_state.last_login = "로그인 이력 없음"

VALID_ID = "shjeon"
VALID_PW = "@jsh2143033"

# -----------------------------
# 1. 기본 상수
# -----------------------------
RI = {1: 0, 2: 0, 3: 0.58, 4: 0.9, 5: 1.12, 6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49}

FUZZY_SCALE = {
    1: (1, 1, 1),
    2: (1, 2, 3),
    3: (2, 3, 4),
    4: (3, 4, 5),
    5: (4, 5, 6),
    6: (5, 6, 7),
    7: (6, 7, 8),
    8: (7, 8, 9),
    9: (9, 9, 9),
}

# -----------------------------
# 2. AHP 관련 함수
# -----------------------------
def convert_punch_to_matrix(punch_data, n_factors):
    mat = np.ones((n_factors, n_factors))
    idx = 0
    for i in range(n_factors):
        for j in range(i + 1, n_factors):
            v = punch_data[idx]
            if v < 0:
                a = abs(v)
                if a > 1:
                    mat[i, j] = 1 / a
                    mat[j, i] = a
            elif v > 1:
                mat[i, j] = v
                mat[j, i] = 1 / v
            idx += 1
    return mat


def ahp_weights_geometric(matrix):
    n = matrix.shape[0]
    gm_row = np.prod(matrix, axis=1) ** (1.0 / n)
    w = gm_row / gm_row.sum()

    eigvals, _ = linalg.eig(matrix)
    lam_max = np.max(eigvals.real)
    CI = (lam_max - n) / (n - 1) if n > 1 else 0
    CR = CI / RI.get(n, 1.49) if n > 2 else 0
    return w, lam_max, CI, CR


def correct_matrix(matrix, threshold=0.1, max_iter=20, alpha=0.3):
    mat = matrix.astype(float).copy()
    w, lam, CI, CR = ahp_weights_geometric(mat)
    orig_CR = CR
    it = 0

    if CR <= threshold:
        return mat, orig_CR, CR, it

    n = mat.shape[0]
    while CR > threshold and it < max_iter:
        w, _, _, _ = ahp_weights_geometric(mat)
        ideal = np.ones_like(mat)
        for i in range(n):
            for j in range(n):
                ideal[i, j] = w[i] / w[j]

        for i in range(n):
            for j in range(i + 1, n):
                a_ij = mat[i, j]
                ideal_ij = ideal[i, j]
                if a_ij <= 0:
                    a_ij = 1.0
                if ideal_ij <= 0:
                    ideal_ij = 1.0
                log_a = np.log(a_ij)
                log_ideal = np.log(ideal_ij)
                log_new = (1 - alpha) * log_a + alpha * log_ideal
                new_ij = np.exp(log_new)
                mat[i, j] = new_ij
                mat[j, i] = 1.0 / new_ij

        _, _, _, CR = ahp_weights_geometric(mat)
        it += 1
        if CR <= threshold:
            break

    return mat, orig_CR, CR, it


def geometric_mean_matrix(mats):
    if len(mats) == 0:
        return None
    mats = np.array(mats)
    logm = np.log(mats)
    gm = np.exp(logm.mean(axis=0))
    return gm


# -----------------------------
# 3. Fuzzy 연산 함수
# -----------------------------
def saaty_to_fuzzy_scalar(v):
    v = max(1, min(9, int(round(v))))
    return FUZZY_SCALE[v]


def reciprocal_fuzzy(tfn):
    l, m, u = tfn
    return (1 / u, 1 / m, 1 / l)


def fuzzy_add(f1, f2):
    l1, m1, u1 = f1
    l2, m2, u2 = f2
    return (l1 + l2, m1 + m2, u1 + u2)


def defuzzify_tfn_array(Si, method="geometric"):
    L = Si[:, 0]
    M = Si[:, 1]
    U = Si[:, 2]
    if method == "weighted":
        c = (L + 2 * M + U) / 4
    elif method == "arithmetic":
        c = (L + M + U) / 3
    elif method == "geometric":
        L2 = np.where(L <= 0, 1e-9, L)
        M2 = np.where(M <= 0, 1e-9, M)
        U2 = np.where(U <= 0, 1e-9, U)
        c = (L2 * M2 * U2) ** (1 / 3)
    else:
        c = M.copy()
    return c


# -----------------------------
# 4. 개선된 Chang Extent Fuzzy AHP
# -----------------------------
def degree_of_possibility(si, sj):
    l1, m1, u1 = si
    l2, m2, u2 = sj
    if m1 >= m2 and l1 >= l2:
        return 1.0
    if u1 <= l2:
        return 0.0
    return max(0.0, min(1.0, (u1 - l2) / ((u1 - m1) + (m2 - l2))))


def fuzzy_ahp_chang_improved(matrix, defuzzy_method="geometric"):
    n = matrix.shape[0]

    F = np.empty((n, n, 3), dtype=float)
    for i in range(n):
        for j in range(n):
            if i == j:
                F[i, j] = (1, 1, 1)
            else:
                v = matrix[i, j]
                if v >= 1:
                    F[i, j] = saaty_to_fuzzy_scalar(v)
                else:
                    inv = 1 / v
                    F[i, j] = reciprocal_fuzzy(saaty_to_fuzzy_scalar(inv))

    row_sum = np.zeros((n, 3))
    for i in range(n):
        s = (0.0, 0.0, 0.0)
        for j in range(n):
            s = fuzzy_add(s, tuple(F[i, j]))
        row_sum[i] = s

    total = row_sum.sum(axis=0)
    total_l, total_m, total_u = total

    Si = np.zeros((n, 3))
    for i in range(n):
        l_i, m_i, u_i = row_sum[i]
        Si[i, 0] = l_i / total_u
        Si[i, 1] = m_i / total_m
        Si[i, 2] = u_i / total_l

    V = np.ones((n, n))
    for i in range(n):
        for j in range(n):
            if i == j:
                V[i, j] = 1.0
            else:
                V[i, j] = degree_of_possibility(tuple(Si[i]), tuple(Si[j]))

    d = np.ones(n)
    for i in range(n):
        for j in range(n):
            if i == j:
                continue
            d[i] *= V[i, j]

    if d.sum() == 0:
        w_fuzzy = np.ones(n) / n
    else:
        w_fuzzy = d / d.sum()

    crisp_S = defuzzify_tfn_array(Si, method=defuzzy_method)

    return Si, d, w_fuzzy, crisp_S, V


# -----------------------------
# 5. 통계 검정 함수 (요인/그룹)
# -----------------------------
def test_factor_significance(weights_matrix, p_threshold=0.05):
    n_experts, n_factors = weights_matrix.shape
    if n_factors < 2:
        return {
            "method": "none",
            "stat": np.nan,
            "pvalue": np.nan,
            "n_experts": n_experts,
            "n_factors": n_factors,
            "comment": "요인이 2개 미만이므로 통계 검정 불가",
        }

    if n_factors == 2:
        stat, pval = stats.ttest_rel(weights_matrix[:, 0], weights_matrix[:, 1])
        method = "paired_t_test"
    else:
        args = [weights_matrix[:, j] for j in range(n_factors)]
        stat, pval = stats.friedmanchisquare(*args)
        method = "friedman_test"

    return {
        "method": method,
        "stat": stat,
        "pvalue": pval,
        "n_experts": n_experts,
        "n_factors": n_factors,
        "p_threshold": p_threshold,
        "significant": "유의" if pval <= p_threshold else "비유의",
    }


def test_group_significance(all_results, groups, labels_kr, p_threshold=0.05):
    rows = []
    if len(groups) < 2:
        return pd.DataFrame(
            [{"요인": "전체", "method": "none", "stat": np.nan, "pvalue": np.nan,
              "p_threshold": p_threshold, "significant": "그룹 2개 미만"}]
        )

    for fi, lab in enumerate(labels_kr):
        samples = []
        for g in groups:
            w = all_results[g]["w_fuzzy"][fi]
            samples.append(w)
        try:
            stat, pval = stats.f_oneway(*[[w] for w in samples])
        except Exception:
            stat, pval = np.nan, np.nan
        rows.append(
            {
                "요인": lab,
                "method": "oneway_anova",
                "stat": stat,
                "pvalue": pval,
                "p_threshold": p_threshold,
                "significant": "유의" if (pd.notna(pval) and pval <= p_threshold) else "비유의",
            }
        )
    return pd.DataFrame(rows)


# -----------------------------
# 6. 삼각 퍼지 멤버십 함수 (그래프용)
# -----------------------------
def triangular_membership(x, a, b, c):
    """표준 삼각 퍼지 멤버십 함수 a <= b <= c 가 되도록 정렬."""
    a, b, c = sorted([a, b, c])
    y = np.zeros_like(x, dtype=float)

    if c == a:
        return y

    if b > a:
        idx = (x > a) & (x < b)
        y[idx] = (x[idx] - a) / (b - a)

    y[(x == b)] = 1.0

    if c > b:
        idx = (x > b) & (x < c)
        y[idx] = (c - x[idx]) / (c - b)

    y[(x <= a) | (x >= c)] = 0.0
    return y


# -----------------------------
# 7. 로그인 UI
# -----------------------------
with st.sidebar:
    st.subheader("🔐 로그인")

    if st.session_state.logged_in:
        st.success(f"로그인 완료: {VALID_ID}")
        st.write(f"최근 로그인 일자: {st.session_state.last_login}")
        if st.button("로그아웃"):
            st.session_state.logged_in = False
    else:
        login_id = st.text_input("아이디", value="", key="login_id")
        login_pw = st.text_input("비밀번호", value="", type="password", key="login_pw")
        if st.button("로그인"):
            if (login_id == VALID_ID) and (login_pw == VALID_PW):
                st.session_state.logged_in = True
                st.session_state.last_login = datetime.now().strftime("%Y-%m-%d %H:%M")
                st.success("로그인 성공")
            else:
                st.error("아이디 또는 비밀번호가 올바르지 않습니다.")
        st.write(f"최근 로그인 일자: {st.session_state.last_login}")

if not st.session_state.logged_in:
    st.title("📊 Fuzzy AHP 분석 시스템")
    st.markdown("제작: 전상현 / jeon080423@gmail.com")
    st.warning("좌측 로그인 후에만 분석 기능을 사용할 수 있습니다.")
    st.stop()

# -----------------------------
# 8. 메인 UI
# -----------------------------
st.title("📊 Fuzzy AHP 분석 시스템")
st.markdown("제작: 전상현 / jeon080423@gmail.com")
st.markdown("AHP와 Fuzzy AHP를 동시에 분석하는 웹 기반 도구.")

with st.sidebar:
    st.header("⚙️ 분석 옵션")

    options = [
        "기하평균 ((l×m×u)^(1/3))",
        "산술평균 ((l+m+u)/3)",
        "가중평균 ((l+2m+u)/4)",
    ]
    defuzz_disp = st.selectbox("비퍼지화 방법 (Si 비퍼지화)", options)
    defuzz_map = {
        "기하평균 ((l×m×u)^(1/3))": "geometric",
        "산술평균 ((l+m+u)/3)": "arithmetic",
        "가중평균 ((l+2m+u)/4)": "weighted",
    }
    defuzz_method = defuzz_map[defuzz_disp]

    cr_th = st.slider("CR 허용 임계값", 0.0, 0.2, 0.1, 0.01)
    p_ttest_threshold = st.number_input(
        "모형간 t-검정 기준 p-value", 0.0, 1.0, 0.05, 0.01, format="%.2f"
    )
    p_factor_threshold = st.number_input(
        "요인간 유의성 기준 p-value", 0.0, 1.0, 0.05, 0.01, format="%.2f"
    )
    p_group_threshold = st.number_input(
        "그룹간 유의성 기준 p-value", 0.0, 1.0, 0.05, 0.01, format="%.2f"
    )

# -----------------------------
# 9. 샘플 & 업로드
# -----------------------------
st.markdown("### 📥 샘플 데이터 (1_2 형식 예시)")
sample_df = pd.DataFrame(
    {
        "ID": [1, 2, 3, 4, 5, 6],
        "type": [1, 1, 1, 1, 1, 1],
        "1_2": [3, 5, 2, -2, -3, -1],
        "1_3": [5, 7, 4, 3, 5, 2],
        "1_4": [7, 9, 5, 5, 7, 4],
        "2_3": [3, 5, 3, 5, 7, 4],
        "2_4": [5, 7, 4, 7, 9, 6],
        "3_4": [3, 5, 2, 5, 7, 3],
    }
)
buf_sample = io.BytesIO()
with pd.ExcelWriter(buf_sample) as w:
    sample_df.to_excel(w, index=False, sheet_name="Sample")
st.download_button(
    "📄 샘플 다운로드",
    buf_sample.getvalue(),
    "fuzzy_ahp_sample_1_2.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.markdown("### 📤 데이터 업로드")
file = st.file_uploader("Excel 파일을 업로드하세요", type=["xlsx", "xls"])

if not file:
    st.info("👆 Excel 파일을 업로드하면 분석을 시작할 수 있습니다.")
    st.stop()

excel_file = pd.ExcelFile(file)[web:409]
sheet_name_used = excel_file.sheet_names[0]
df = pd.read_excel(excel_file, sheet_name=sheet_name_used)

st.success(f"파일 업로드 완료: {len(df)}행 (시트명: {sheet_name_used})")
with st.expander("📋 데이터 미리보기"):
    st.dataframe(df.head())

id_col = df.columns[0]
type_col = df.columns[1]
comp_cols = df.columns[2:]

n_comp = len(comp_cols)
n_factor = int((1 + np.sqrt(1 + 8 * n_comp)) / 2)

index_set = set()
for c in comp_cols:
    name = str(c)
    if "_" in name:
        a, b = name.split("_")
        index_set.add(int(a))
        index_set.add(int(b))
if len(index_set) == n_factor:
    labels_kr = [f"요인{i}" for i in sorted(index_set)]
else:
    labels_kr = [f"요인{i+1}" for i in range(n_factor)]

labels_en = [f"Factor{i+1}" for i in range(len(labels_kr))]

has_group = df[type_col].notna().any()
groups = df[type_col].dropna().unique() if has_group else ["All"]

# -----------------------------
# 10. 분석 실행
# -----------------------------
if st.button("🚀 분석 시작", type="primary"):
    all_results = {}
    cons_list = []
    factor_tests = []
    fuzzy_raw_rows = []
    ahp_result_rows = []
    fuzzy_result_rows = []
    compare_all_rows = []
    raw_data_df = df.copy()

    prog = st.progress(0.0)
    step = 1.0 / len(groups)

    for gi, g in enumerate(groups):
        gdf = df[df[type_col] == g] if has_group else df

        matrices = []
        for _, row in gdf.iterrows():
            punch = pd.to_numeric(row[comp_cols], errors="coerce").fillna(1).values
            mat = convert_punch_to_matrix(punch, n_factor)
            cmat, cr0, cr1, it = correct_matrix(mat, threshold=cr_th, max_iter=20, alpha=0.3)

            cons_list.append(
                {
                    "ID": row[id_col],
                    "Group": g if has_group else "All",
                    "보정 전 CR": cr0,
                    "보정 후 CR": cr1,
                    "보정 횟수": it,
                    "일관성": "○" if cr1 <= cr_th else "×",
                }
            )
            matrices.append(cmat)

            Si_i, d_i, w_fuzzy_i, crisp_S_i, V_i = fuzzy_ahp_chang_improved(cmat, defuzz_method)
            row_dict = {"ID": row[id_col], "Group": g if has_group else "All"}
            for fi, lab in enumerate(labels_kr):
                row_dict[f"{lab}_Lower"] = Si_i[fi, 0]
                row_dict[f"{lab}_Medium"] = Si_i[fi, 1]
                row_dict[f"{lab}_Upper"] = Si_i[fi, 2]
                row_dict[f"{lab}_Norm"] = w_fuzzy_i[fi]
            fuzzy_raw_rows.append(row_dict)

        gm = geometric_mean_matrix(matrices)
        w_ahp, lam, CI, CR = ahp_weights_geometric(gm)
        Si, d_raw, w_fuzzy, crisp_S, V = fuzzy_ahp_chang_improved(gm, defuzz_method)

        fuzzy_matrix = np.ones_like(gm)
        for i in range(n_factor):
            for j in range(n_factor):
                fuzzy_matrix[i, j] = w_fuzzy[i] / w_fuzzy[j]

        all_results[g] = {
            "matrix": gm,
            "fuzzy_matrix": fuzzy_matrix,
            "ahp_w": w_ahp,
            "lam": lam,
            "CI": CI,
            "CR": CR,
            "Si": Si,
            "d_raw": d_raw,
            "w_fuzzy": w_fuzzy,
            "crisp_S": crisp_S,
            "V": V,
        }

        ahp_rank = pd.Series(w_ahp).rank(ascending=False, method="min").astype(int)
        for fi, lab in enumerate(labels_kr):
            ahp_result_rows.append(
                {
                    "그룹": g if has_group else "All",
                    "요인": lab,
                    "AHP_가중치": w_ahp[fi],
                    "AHP_순위": int(ahp_rank[fi]),
                    "lambda_max": lam,
                    "CI": CI,
                    "CR": CR,
                }
            )

        fuzzy_rank = pd.Series(w_fuzzy).rank(ascending=False, method="min").astype(int)
        for fi, lab in enumerate(labels_kr):
            fuzzy_result_rows.append(
                {
                    "그룹": g if has_group else "All",
                    "요인": lab,
                    "Si_Lower": Si[fi, 0],
                    "Si_Medium": Si[fi, 1],
                    "Si_Upper": Si[fi, 2],
                    "Crisp_Si": crisp_S[fi],
                    "d_i": d_raw[fi],
                    "Fuzzy_가중치": w_fuzzy[fi],
                    "Fuzzy_순위": int(fuzzy_rank[fi]),
                }
            )

        diff_rank = fuzzy_rank - ahp_rank
        for fi, lab in enumerate(labels_kr):
            compare_all_rows.append(
                {
                    "그룹": g if has_group else "All",
                    "요인": lab,
                    "AHP_가중치": w_ahp[fi],
                    "AHP_순위": int(ahp_rank[fi]),
                    "Fuzzy_가중치": w_fuzzy[fi],
                    "Fuzzy_순위": int(fuzzy_rank[fi]),
                    "순위변동": int(diff_rank[fi]),
                }
            )

        weights_mat = np.tile(w_fuzzy, (len(gdf), 1))
        test_res = test_factor_significance(weights_mat, p_threshold=p_factor_threshold)
        test_res["Group"] = g
        factor_tests.append(test_res)

        prog.progress((gi + 1) * step)

    st.success("분석 완료")

    cons_df = pd.DataFrame(cons_list)
    factor_test_df = pd.DataFrame(factor_tests)
    fuzzy_raw_df = pd.DataFrame(fuzzy_raw_rows)
    ahp_result_df = pd.DataFrame(ahp_result_rows)
    fuzzy_result_df = pd.DataFrame(fuzzy_result_rows)
    compare_all_df = pd.DataFrame(compare_all_rows)
    group_effect_df = test_group_significance(all_results, groups, labels_kr, p_threshold=p_group_threshold)

    fmt3 = "{:.3f}"

    def style3(df, cols=None):
        if cols is None:
            return df.style.format(fmt3)
        return df.style.format({c: fmt3 for c in cols})

    tabs = st.tabs(
        [
            "일관성 검증",
            "AHP/Fuzzy 행렬 + TFN",
            "AHP/Fuzzy 결과",
            "요인/그룹 유의성",
            "엑셀 저장",
        ]
    )

    # ---------------- 표시 탭 ----------------
    with tabs[0]:
        st.dataframe(style3(cons_df, cons_df.select_dtypes("number").columns), use_container_width=True)

    with tabs[1]:
        for g, r in all_results.items():
            st.markdown(f"### 그룹: {g}")
            mat_df = pd.DataFrame(r["matrix"], index=labels_kr, columns=labels_kr)
            fuzzy_mat_df = pd.DataFrame(r["fuzzy_matrix"], index=labels_kr, columns=labels_kr)

            st.subheader("일반 AHP 최종 판단행렬")
            st.dataframe(style3(mat_df), use_container_width=True)
            st.subheader("Fuzzy AHP 최종 판단행렬")
            st.dataframe(style3(fuzzy_mat_df), use_container_width=True)

            # ---- 삼각퍼지 그래프 (한 그래프에 모든 요인, 범례 영어) ----
            st.subheader("Triangular Fuzzy Numbers (All Factors)")
            Si = r["Si"]

            a_list, b_list, c_list = [], [], []
            for fi in range(len(labels_kr)):
                l, m, u = Si[fi]
                a, b, c = sorted([float(l), float(m), float(u)])
                a_list.append(a)
                b_list.append(b)
                c_list.append(c)

            global_a = min(a_list)
            global_c = max(c_list)
            x = np.linspace(global_a, global_c, 400)

            fig, ax = plt.subplots()
            for fi, lab_en in enumerate(labels_en):
                a, b, c = sorted([a_list[fi], b_list[fi], c_list[fi]])
                if c == a:
                    continue
                y = triangular_membership(x, a, b, c)
                ax.plot(x, y, label=lab_en)

            ax.set_title(f"Triangular Fuzzy Numbers (Group: {g})")
            ax.set_xlabel("Value")
            ax.set_ylabel("Membership")
            ax.set_ylim(0, 1.05)
            ax.grid(True, alpha=0.3)
            ax.legend(loc="best")
            st.pyplot(fig)

    with tabs[2]:
        st.subheader("AHP 결과")
        st.dataframe(
            ahp_result_df.style.format(
                {
                    "AHP_가중치": fmt3,
                    "lambda_max": fmt3,
                    "CI": fmt3,
                    "CR": fmt3,
                }
            ),
            use_container_width=True,
        )
        st.subheader("Fuzzy 결과")
        st.dataframe(
            fuzzy_result_df.style.format(
                {
                    "Si_Lower": fmt3,
                    "Si_Medium": fmt3,
                    "Si_Upper": fmt3,
                    "Crisp_Si": fmt3,
                    "d_i": fmt3,
                    "Fuzzy_가중치": fmt3,
                }
            ),
            use_container_width=True,
        )
        st.subheader("비교_All (AHP vs Fuzzy)")
        st.dataframe(
            compare_all_df.style.format(
                {"AHP_가중치": fmt3, "Fuzzy_가중치": fmt3}
            ),
            use_container_width=True,
        )

    with tabs[3]:
        st.subheader("요인간 유의성 (그룹 내부)")
        st.dataframe(style3(factor_test_df, factor_test_df.select_dtypes("number").columns), use_container_width=True)
        st.subheader("그룹간 유의성 (요인별)")
        st.dataframe(style3(group_effect_df, group_effect_df.select_dtypes("number").columns), use_container_width=True)

    # ---------------- 엑셀 저장 ----------------
    with tabs[4]:
        st.markdown("### 📊 분석 결과 엑셀 저장")

        def apply_number_format_000(ws):
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.000"

        def create_excel_report():
            out = io.BytesIO()
            with pd.ExcelWriter(out, engine="openpyxl") as writer:
                # 1. FuzzyAHP 로우데이터 / 원본데이터
                fuzzy_raw_df.to_excel(writer, sheet_name="FuzzyAHP_로우데이터", index=False)
                raw_data_df.to_excel(writer, sheet_name="원본데이터", index=False)

                # 2. 일관성 검증
                cons_df.to_excel(writer, sheet_name="일관성검증", index=False)

                # 3. 행렬_All (AHP + Fuzzy 위/아래)
                g0 = list(all_results.keys())[0]
                r0 = all_results[g0]
                ahp_mat = pd.DataFrame(r0["matrix"], index=labels_kr, columns=labels_kr)
                fuzzy_mat = pd.DataFrame(r0["fuzzy_matrix"], index=labels_kr, columns=labels_kr)

                block_top = ahp_mat.copy()
                block_top.insert(0, "구분", labels_kr)

                block_bottom = fuzzy_mat.copy()
                block_bottom.insert(0, "구분", labels_kr)

                blank = pd.DataFrame([[""] * block_top.shape[1]])

                out_mat = pd.concat(
                    [
                        pd.DataFrame(
                            [["일반 AHP 최종 판단행렬 (Group: All)"] + [""] * (block_top.shape[1] - 1)]
                        ),
                        block_top.reset_index(drop=True),
                        blank,
                        pd.DataFrame(
                            [["Fuzzy AHP 최종 판단행렬 (Group: All)"] + [""] * (block_bottom.shape[1] - 1)]
                        ),
                        block_bottom.reset_index(drop=True),
                    ],
                    ignore_index=True,
                )
                out_mat.to_excel(writer, sheet_name="행렬_All", index=False, header=False)

                # 4. AHP/Fuzzy/비교 결과
                ahp_result_df.to_excel(writer, sheet_name="AHP결과", index=False)
                fuzzy_result_df.to_excel(writer, sheet_name="Fuzzy결과", index=False)
                compare_all_df.to_excel(writer, sheet_name="비교_All", index=False)

                # 5. 요인간 / 그룹간 유의성
                factor_test_df.to_excel(writer, sheet_name="요인간_유의성", index=False)
                group_effect_df.to_excel(writer, sheet_name="그룹간_유의성", index=False)

                # 6. 분석 설정
                setting_df = pd.DataFrame(
                    {
                        "설정항목": [
                            "비퍼지화_방법",
                            "CR_임계값",
                            "t검정_p기준",
                            "요인간_p기준",
                            "그룹간_p기준",
                        ],
                        "값": [
                            defuzz_method,
                            cr_th,
                            p_ttest_threshold,
                            p_factor_threshold,
                            p_group_threshold,
                        ],
                    }
                )
                setting_df.to_excel(writer, sheet_name="분석설정", index=False)

                # ---------- openpyxl 객체에 접근하여 포맷팅/차트 ----------
                wb = writer.book

                for sheet_name in [
                    "FuzzyAHP_로우데이터",
                    "원본데이터",
                    "일관성검증",
                    "행렬_All",
                    "AHP결과",
                    "Fuzzy결과",
                    "비교_All",
                    "요인간_유의성",
                    "그룹간_유의성",
                ]:
                    ws = wb[sheet_name]
                    apply_number_format_000(ws)

                # Fuzzy TFN 그래프용 데이터 + 차트 시트 (모든 요인을 한 그래프에, 범례 영어)[web:371]
                chart_sheet = wb.create_sheet("Fuzzy_그래프_시트")
                chart_sheet.append(["x"] + labels_en)

                first_group = list(all_results.keys())[0]
                Si0 = all_results[first_group]["Si"]

                abc_list = []
                for fi in range(len(labels_en)):
                    l, m, u = Si0[fi]
                    a, b, c = sorted([float(l), float(m), float(u)])
                    abc_list.append((a, b, c))
                global_a = min(a for a, b, c in abc_list)
                global_c = max(c for a, b, c in abc_list)

                x_vals = np.linspace(global_a, global_c, 50)

                for xv in x_vals:
                    row_vals = [float(xv)]
                    for (a, b, c) in abc_list:
                        yv = float(triangular_membership(np.array([xv]), a, b, c)[0])
                        row_vals.append(yv)
                    chart_sheet.append(row_vals)

                chart = LineChart()
                chart.title = "Triangular Fuzzy Numbers (All Factors)"
                chart.y_axis.title = "Membership"
                chart.x_axis.title = "Value"

                data = Reference(
                    chart_sheet,
                    min_col=2,
                    min_row=1,
                    max_col=1 + len(labels_en),
                    max_row=1 + len(x_vals),
                )
                cats = Reference(
                    chart_sheet,
                    min_col=1,
                    min_row=2,
                    max_row=1 + len(x_vals),
                )
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart_sheet.add_chart(chart, "H2")

            out.seek(0)
            return out.getvalue()

        excel_bytes = create_excel_report()
        out_name = f"{sheet_name_used}_FAHP_분석결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button(
            "📥 분석 결과 다운로드 (Excel)",
            data=excel_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )
